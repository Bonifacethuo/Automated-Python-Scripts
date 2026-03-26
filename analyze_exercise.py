import pandas as pd
import os

directory = r"c:\Users\USER\Desktop\China"

with open(os.path.join(directory, "output8.txt"), "w", encoding="utf-8") as f:
    f.write("--- Exercise 2.2 ---\n")
    excel_file = os.path.join(directory, "Exercise 2.2 (1).xlsx")
    df = pd.read_excel(excel_file, sheet_name='Data')
    f.write("Columns: " + str(df.columns.tolist()) + "\n")
    f.write(df.head(5).to_string() + "\n")

    f.write("\n--- Exercise 2.2 Solution ---\n")
    sol_file = os.path.join(directory, "Exercise 2.2 Solution (1).xlsx")
    # check available sheets
    import openpyxl
    wb = openpyxl.load_workbook(sol_file)
    f.write("Solution Sheets: " + str(wb.sheetnames) + "\n")
    
    # Try reading the first sheet
    sol_df = pd.read_excel(sol_file, sheet_name=wb.sheetnames[0])
    f.write("Columns: " + str(sol_df.columns.tolist()) + "\n")
    f.write(sol_df.head(15).to_string() + "\n")
